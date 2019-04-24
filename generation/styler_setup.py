from generation.random_physicalculture import Data_Generate
from openpyxl.styles import Font, Alignment


class A_Column:
    def column(self):
        pass


class B_Column:
    def column(self, ws):
        for row in range(19, 25):
            _ = ws.cell(column=2, row=row, value="{0}".format(Data_Generate().gen_timesleep()))
            print(Data_Generate().gen_timesleep())

        for cell in ws['B19:B25']:
            cell[0].font = Font(name='Times New Roman',
                                size=14,
                                color='FF0000')

            cell[0].alignment = Alignment(horizontal='center',
                                          vertical='center')


class C_Column:
    def column(self, ws):
        for row in range(19, 25):
            _ = ws.cell(column=3, row=row, value="{0}".format(Data_Generate().gen_appetite()))

        for cell in ws['C19:C25']:
            cell[0].font = Font(name='Times New Roman',
                                size=14,
                                color='FF0000')

            cell[0].alignment = Alignment(horizontal='center',
                                          vertical='center')


class D_Column:
    def column(self,ws):

        for row in range(19, 25):
            _ = ws.cell(column=4, row=row, value="{0}".format(Data_Generate.gen_pulse()))
            print(_)

        for cell in ws['D19:D25']:
            cell[0].font = Font(name='Times New Roman',
                                size=14,
                                color='FF0000')

            cell[0].alignment = Alignment(horizontal='center',
                                          vertical='center')


class E_Column:
    def column(self, ws):
        for row in range(19, 25):
            _ = ws.cell(column=5, row=row, value="{0}".format(Data_Generate.gen_weight()))

        for cell in ws['E19:E25']:
            cell[0].font = Font(name='Times New Roman',
                                size=14,
                                color='FF0000')

            cell[0].alignment = Alignment(horizontal='center',
                                          vertical='center')


class F_Column:
    def column(self, ws):
        for row in range(19, 25):
            _ = ws.cell(column=6, row=row, value="{0}".format(Data_Generate.gen_gymnastic()))
            print(Data_Generate.gen_gymnastic())

        for cell in ws['F19:F25']:
            cell[0].font = Font(name='Times New Roman',
                                size=14,
                                color='FF0000')

            cell[0].alignment = Alignment(horizontal='center',
                                          vertical='center')


class G_AND_H_Column:
    def column(self, ws):
        list_name_physical_culture = []
        for i, row in enumerate(range(19, 25)):
            _ = ws.cell(column=7, row=row, value="{0}".format(Data_Generate.gen_name_physical_culture()))
            list_name_physical_culture.append(_.value)

        for cell in ws['G19:G26']:
            cell[0].font = Font(name='Times New Roman',
                                size=14,
                                color='FF0000')

            cell[0].alignment = Alignment(horizontal='center',
                                          vertical='center')

        for i, row in enumerate(range(19, 25)):
            _ = ws.cell(column=8, row=row, value="{0}".format(Data_Generate().gen_counts_exercises(list_name_physical_culture[i])))

        for cell in ws['H19:H26']:
            cell[0].font = Font(name='Times New Roman',
                                size=14,
                                color='FF0000')

            cell[0].alignment = Alignment(horizontal='center',
                                          vertical='center')

        return list_name_physical_culture






