from openpyxl import load_workbook
from generation.data_filling import *


def Header(ws):
    surname = input('Введи свою фамилию\n')
    name = input('Введи своё имя\n')
    last_name = input('Введи своё отчество\n')
    study_group = input('В какой группе учишься?\n')
    date_birthday = input('Укажи дату своего рождения родился?\nПример: 18/09/99\n')

    ws['I9'] = '{0} {1} {2}' .format(surname, name, last_name)
    ws['I10'] = '{0}'.format(study_group)
    ws['I11'] = '{0}'.format(date_birthday)


def columns_fill(ws, count_days_generate, personal_weight):

    first_column = A_Column(ws, count_days_generate)
    second_column = B_Column(ws, count_days_generate)
    three_column = C_Column(ws, count_days_generate)
    four_column = D_Column(ws, count_days_generate)
    fifth_column = E_Column(ws, count_days_generate, personal_weight)
    sixth_column = F_Column(ws, count_days_generate)
    seventh_column = G_AND_H_AND_I_Column(ws, count_days_generate)
    eighth_column = J_Column(ws, count_days_generate)


if __name__ == '__main__':
    print('Welcome to the GenCulture')

    wb = load_workbook(filename='download.xlsx')

    ws = wb.active

    Header(ws)

    personal_weight = int(input('Сколько ты весишь?\nПример: 75\n'))

    count_days_generate = int(input('Сколько дней хочешь сгенерировать?\n Рекомендуется -> 50\n'))

    columns_fill(ws, count_days_generate, personal_weight)

    wb.save('physical_culture.xlsx')
